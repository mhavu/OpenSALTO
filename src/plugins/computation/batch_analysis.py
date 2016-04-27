#
#  batch_analysis.py
#  OpenSALTO
#
#  An OpenSALTO batch analysis report plugin
#
#  Copyright 2016 Marko Havu. Released under the terms of
#  GNU General Public License version 3 or later.
#

import salto
import os, datetime, math
import csv, warnings
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Style, Font

def setStyle(ws, value, style):
    # Silence the warning about using a composite Style object instead of Font.
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        try:
            result = [Cell(ws, column='A', row = 1, value=item) for item in value]
            for cell in result:
                cell.style = style
        except TypeError:
            result = Cell(ws, column='A', row = 1, value=value)
            result.style = style
    return result

class Plugin(salto.Plugin):
    """OpenSALTO plugin for batch analysis
        
        Expects the analysis plugin to return a report in the result dictionary
        with key html (for html reports), or table (for table data). Table
        reports should be dictionaries with keys header,  data, and summary for
        a header row, data rows, and summary rows, respectively. The summary is
        ignored if a flat table format is specified."""
    def __init__(self, manager):
        super(Plugin, self).__init__(manager)
        inputs = [('channelTable', 'S', 0, 0),
                  ('source', 'O', "list of data files or directories", None),
                  ('source_format', 'S', "name of the file format plugin to be used", 'directory'),
                  ('result_file_path', 'S', "path of the result file", None),
                  ('result_format', 'S', "format of the result file", 'xlsx'),
                  ('analysis', 'S', "name of the analysis plugin to be used", None),
                  ('parameters', 'f', "input parameters for the analysis plugin", {}),
                  ('options', 'O', "sequence of options ('flat', 'structured')", ('flat')),
                  ('communicator', 'O', "MPI communicator object (optional)", None)
                  ]
        self.registerComputation('batch analysis', self._report,
                                 inputs = inputs,
                                 outputs = [])
        self.datestyle = Style(number_format='yyyy-mm-dd h:mm:ss')
        self.boldstyle = Style(font=Font(bold=True))

    def _report(self, inputs):
        comm = inputs.get('communicator')
        if comm:
            # Divide work between MPI nodes.
            size = comm.Get_size()
            rank = comm.Get_rank()
            if rank == 0:
                each = math.ceil(len(inputs['source']) / size)
                source = [inputs['source'][i*each:(i+1)*each] for i in range(size)]
                source.reverse()
            else:
                source = None
            source = comm.scatter(source, root=0)
        else:
            rank = 0
            source = inputs['source']
        isFlat = 'flat' in inputs['options']
        # Check options.
        if (not isFlat) != ('structured' in inputs['options']):
            raise ValueException("Options 'flat' and 'structured' are mutually exclusive")
        # Check the that filenames in source are unique.
        basenames = [os.path.basename(path) for path in source]
        if len(basenames) != len(set(basenames)):
            raise ValueException('The basenames in source are not unique')
        # Run the analysis.
        chTables = salto.channelTables
        pm = salto.pluginManager
        tableName = salto.makeUniqueKey(salto.channelTables, 'temp')
        reports = {}
        for path, basename in zip(source, basenames):
            print(basename)
            chTables[tableName] = salto.ChannelTable()
            pm.read(path, inputs['source_format'], tableName)
            parameters = inputs['parameters'].copy()
            parameters['channelTable'] = tableName
            result = pm.compute(inputs['analysis'], parameters)
            tmpTable = result.get('channelTable')
            if tmpTable:
                chTables.pop(tmpTable, None)
            if inputs['result_format'].lower() == 'html':
                reports[basename] = result['html']
            else:
                reports[basename] = result['table']
        chTables.pop(tableName, None)
        if comm:
            reports = comm.gather(reports, root=0)
            if rank == 0:
                reports = {k: v for d in reports for k, v in d.items()}
        # Create the report.
        if not comm or rank == 0:
            datestr = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            params = pm.computations[inputs['analysis']].computations[inputs['analysis']][1]
            preamble = [["analysis date", datestr],
                        ["analysis plugin", inputs['analysis']]]
            preamble += [[p[0], inputs.get(p[0], p[3]), p[2]]
                         for p in params if p[0] != 'channelTable']
            if inputs['result_format'].lower() == 'xlsx':
                wb = Workbook()
                ws = wb.active
                ws.title = "settings"
                for row in preamble:
                    ws.append(row)
                # TODO: Set column widths.
                if isFlat:
                    ws = wb.create_sheet()
                    ws.title = "results"
                    headers = [reports[k]['header'] for k in reports
                               if 'header' in reports[k]]
                    if headers and headers.count(headers[0]) == len(headers):
                        # All headers are equal.
                        ws.append(setStyle(ws, ['Filename'] + headers[0], self.boldstyle))
                    else:
                        raise NotImplementedError("Combining tables with different structures is not implemented yet")
                    for basename, report in reports.items():
                        for row in report['data']:
                            ws.append([basename] + row)
                    # TODO: Set column widths.
                else:
                    for basename, report in reports.items():
                        ws = wb.create_sheet()
                        ws.title = basename
                        if 'header' in report:
                            ws.append(setStyle(ws, report['header'], self.boldstyle))
                        for row in report['data']:
                            ws.append(row)
                        for row in report.get('summary', []):
                            ws.append(setStyle(ws, row, self.boldstyle))
                        # TODO: Set column widths.
                wb.save(inputs['result_file_path'])
            elif inputs['result_format'].lower() == 'csv':
                raise NotImplementedError("CSV reporting is not implemented yet")
            elif inputs['result_format'].lower() == 'html':
                raise NotImplementedError("HTML reporting is not implemented yet")
        return {}
